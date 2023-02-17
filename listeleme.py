import os

import openpyxl
from PyQt5 import QtCore, QtGui, QtWidgets
import sqlite3

from openpyxl.workbook import Workbook

baglanti = sqlite3.connect(".\Data\stok.db")
islem = baglanti.cursor()

class Ui_Form_listele(object):
    def setupUi(self, Form):
        Form.setObjectName("Form")
        Form.resize(1880, 794)
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap(":/newPrefix/Stok/note.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        Form.setWindowIcon(icon)
        self.widget = QtWidgets.QWidget(Form)
        self.widget.setGeometry(QtCore.QRect(0, 0, 1951, 871))
        self.widget.setStyleSheet("#widget{\n"
" background-color: #3DD1E7;}")
        self.widget.setObjectName("widget")
        self.tableWidget = QtWidgets.QTableWidget(self.widget)
        self.tableWidget.setGeometry(QtCore.QRect(10, 50, 1451, 731))
        self.tableWidget.setMaximumSize(QtCore.QSize(16777190, 16777215))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.tableWidget.setFont(font)
        self.tableWidget.setRowCount(1000)
        self.tableWidget.setColumnCount(11)
        self.tableWidget.setObjectName("tableWidget")
        item = QtWidgets.QTableWidgetItem()
        font = QtGui.QFont()
        font.setPointSize(11)
        font.setBold(True)
        font.setWeight(75)
        item.setFont(font)
        self.tableWidget.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        font = QtGui.QFont()
        font.setPointSize(11)
        font.setBold(True)
        font.setWeight(75)
        item.setFont(font)
        self.tableWidget.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        font = QtGui.QFont()
        font.setPointSize(11)
        font.setBold(True)
        font.setWeight(75)
        item.setFont(font)
        self.tableWidget.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        font = QtGui.QFont()
        font.setPointSize(11)
        font.setBold(True)
        font.setWeight(75)
        item.setFont(font)
        self.tableWidget.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        font = QtGui.QFont()
        font.setPointSize(11)
        font.setBold(True)
        font.setWeight(75)
        item.setFont(font)
        self.tableWidget.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        font = QtGui.QFont()
        font.setPointSize(11)
        font.setBold(True)
        font.setWeight(75)
        item.setFont(font)
        self.tableWidget.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        font = QtGui.QFont()
        font.setPointSize(11)
        font.setBold(True)
        font.setWeight(75)
        item.setFont(font)
        self.tableWidget.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        font = QtGui.QFont()
        font.setPointSize(11)
        font.setBold(True)
        font.setWeight(75)
        item.setFont(font)
        self.tableWidget.setHorizontalHeaderItem(7, item)
        item = QtWidgets.QTableWidgetItem()
        font = QtGui.QFont()
        font.setPointSize(11)
        font.setBold(True)
        font.setWeight(75)
        item.setFont(font)
        self.tableWidget.setHorizontalHeaderItem(8, item)
        item = QtWidgets.QTableWidgetItem()
        font = QtGui.QFont()
        font.setPointSize(11)
        font.setBold(True)
        font.setWeight(75)
        item.setFont(font)
        self.tableWidget.setHorizontalHeaderItem(9, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignLeading|QtCore.Qt.AlignBottom)
        font = QtGui.QFont()
        font.setPointSize(11)
        font.setBold(True)
        font.setWeight(75)
        item.setFont(font)
        self.tableWidget.setHorizontalHeaderItem(10, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setItem(0, 10, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignLeading|QtCore.Qt.AlignTop)
        self.tableWidget.setItem(3, 10, item)
        self.label_2 = QtWidgets.QLabel(self.widget)
        self.label_2.setGeometry(QtCore.QRect(1460, 480, 141, 301))
        self.label_2.setStyleSheet("background-color:rgba(0, 0, 0, 80);\n"
"border-bottom-right-radius: 50px;\n"
"\n"
"")
        self.label_2.setText("")
        self.label_2.setObjectName("label_2")
        self.pushButton_9 = QtWidgets.QPushButton(self.widget)
        self.pushButton_9.setGeometry(QtCore.QRect(1460, 480, 141, 61))
        font = QtGui.QFont()
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_9.setFont(font)
        self.pushButton_9.setStyleSheet("QPushButton{\n"
"    background-color: qlineargradient(spread:pad, x1:0, y1:0.505682, x2:1, y2:0.477, stop:0 rgba(11, 131, 120, 219), stop:1 rgba(85, 98, 112, 226));\n"
"    color:rgba(255, 255, 255, 210);\n"
"    border-radius: 50px;\n"
"}\n"
"\n"
"QPushButton:hover{\n"
"    background-color: qlineargradient(spread:pad, x1:0, y1:0.505682, x2:1, y2:0.477, stop:0 rgba(150, 123, 111, 219), stop:1 rgba(85, 81, 84, 226));\n"
"}\n"
"\n"
"QPushButton:pressed{\n"
"    padding-left:5px;\n"
"    padding-top:5px;\n"
"    background-color:rgba(150, 123, 111, 255);\n"
"}\n"
"\n"
"\n"
"\n"
"border-left-radius: 25px;")
        icon1 = QtGui.QIcon()
        icon1.addPixmap(QtGui.QPixmap(":/newPrefix/Stok/trash.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.pushButton_9.setIcon(icon1)
        self.pushButton_9.setIconSize(QtCore.QSize(27, 27))
        self.pushButton_9.setObjectName("pushButton_9")
        self.pushButton_9.clicked.connect(self.satirsil)
        self.pushButton_10 = QtWidgets.QPushButton(self.widget)
        self.pushButton_10.setGeometry(QtCore.QRect(1460, 540, 141, 61))
        font = QtGui.QFont()
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_10.setFont(font)
        self.pushButton_10.setStyleSheet("QPushButton{\n"
"    background-color: qlineargradient(spread:pad, x1:0, y1:0.505682, x2:1, y2:0.477, stop:0 rgba(11, 131, 120, 219), stop:1 rgba(85, 98, 112, 226));\n"
"    color:rgba(255, 255, 255, 210);\n"
"    border-radius: 50px;\n"
"}\n"
"\n"
"QPushButton:hover{\n"
"    background-color: qlineargradient(spread:pad, x1:0, y1:0.505682, x2:1, y2:0.477, stop:0 rgba(150, 123, 111, 219), stop:1 rgba(85, 81, 84, 226));\n"
"}\n"
"\n"
"QPushButton:pressed{\n"
"    padding-left:5px;\n"
"    padding-top:5px;\n"
"    background-color:rgba(150, 123, 111, 255);\n"
"}\n"
"\n"
"\n"
"\n"
"border-left-radius: 25px;")
        self.pushButton_10.setIcon(icon)
        self.pushButton_10.setIconSize(QtCore.QSize(26, 26))
        self.pushButton_10.setObjectName("pushButton_10")
        self.pushButton_11 = QtWidgets.QPushButton(self.widget)
        self.pushButton_11.setGeometry(QtCore.QRect(1460, 720, 141, 61))
        font = QtGui.QFont()
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_11.setFont(font)
        self.pushButton_11.setStyleSheet("QPushButton{\n"
"    background-color: qlineargradient(spread:pad, x1:0, y1:0.505682, x2:1, y2:0.477, stop:0 rgba(11, 131, 120, 219), stop:1 rgba(85, 98, 112, 226));\n"
"    color:rgba(255, 255, 255, 210);\n"
"    border-bottom-right-radius: 50px;\n"
"}\n"
"\n"
"QPushButton:hover{\n"
"    background-color: qlineargradient(spread:pad, x1:0, y1:0.505682, x2:1, y2:0.477, stop:0 rgba(150, 123, 111, 219), stop:1 rgba(85, 81, 84, 226));\n"
"}\n"
"\n"
"QPushButton:pressed{\n"
"    padding-left:5px;\n"
"    padding-top:5px;\n"
"    background-color:rgba(150, 123, 111, 255);\n"
"}\n"
"\n"
"\n"
"\n"
"border-left-radius: 25px;")
        icon2 = QtGui.QIcon()
        icon2.addPixmap(QtGui.QPixmap(":/self/Stok/exit.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.pushButton_11.setIcon(icon2)
        self.pushButton_11.setIconSize(QtCore.QSize(27, 27))
        self.pushButton_11.setObjectName("pushButton_11")
        self.pushButton_12 = QtWidgets.QPushButton(self.widget)
        self.pushButton_12.setGeometry(QtCore.QRect(1460, 600, 141, 61))
        font = QtGui.QFont()
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_12.setFont(font)
        self.pushButton_12.setStyleSheet("QPushButton{\n"
"    background-color: qlineargradient(spread:pad, x1:0, y1:0.505682, x2:1, y2:0.477, stop:0 rgba(11, 131, 120, 219), stop:1 rgba(85, 98, 112, 226));\n"
"    color:rgba(255, 255, 255, 210);\n"
"    border-radius: 50px;\n"
"}\n"
"\n"
"QPushButton:hover{\n"
"    background-color: qlineargradient(spread:pad, x1:0, y1:0.505682, x2:1, y2:0.477, stop:0 rgba(150, 123, 111, 219), stop:1 rgba(85, 81, 84, 226));\n"
"}\n"
"\n"
"QPushButton:pressed{\n"
"    padding-left:5px;\n"
"    padding-top:5px;\n"
"    background-color:rgba(150, 123, 111, 255);\n"
"}\n"
"\n"
"\n"
"\n"
"border-left-radius: 25px;")
        icon3 = QtGui.QIcon()
        icon3.addPixmap(QtGui.QPixmap(":/newPrefix/Stok/printer.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.pushButton_12.setIcon(icon3)
        self.pushButton_12.setIconSize(QtCore.QSize(28, 28))
        self.pushButton_12.setObjectName("pushButton_12")
        self.pushButton_12.clicked.connect(self.print)
        self.pushButton_13 = QtWidgets.QPushButton(self.widget)
        self.pushButton_13.setGeometry(QtCore.QRect(1460, 660, 141, 61))
        font = QtGui.QFont()
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_13.setFont(font)
        self.pushButton_13.setStyleSheet("QPushButton{\n"
"    background-color: qlineargradient(spread:pad, x1:0, y1:0.505682, x2:1, y2:0.477, stop:0 rgba(11, 131, 120, 219), stop:1 rgba(85, 98, 112, 226));\n"
"    color:rgba(255, 255, 255, 210);\n"
"    border-radius: 50px;\n"
"}\n"
"\n"
"QPushButton:hover{\n"
"    background-color: qlineargradient(spread:pad, x1:0, y1:0.505682, x2:1, y2:0.477, stop:0 rgba(150, 123, 111, 219), stop:1 rgba(85, 81, 84, 226));\n"
"}\n"
"\n"
"QPushButton:pressed{\n"
"    padding-left:5px;\n"
"    padding-top:5px;\n"
"    background-color:rgba(150, 123, 111, 255);\n"
"}\n"
"\n"
"\n"
"\n"
"border-left-radius: 25px;")
        icon4 = QtGui.QIcon()
        icon4.addPixmap(QtGui.QPixmap(":/newPrefix/Stok/excel.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.pushButton_13.setIcon(icon4)
        self.pushButton_13.setIconSize(QtCore.QSize(28, 28))
        self.pushButton_13.setObjectName("pushButton_13")
        self.pushButton_13.clicked.connect(self.excel_dok)
        self.label_3 = QtWidgets.QLabel(self.widget)
        self.label_3.setGeometry(QtCore.QRect(1460, 50, 391, 431))
        self.label_3.setStyleSheet("\n"
"background-color: rgb(121, 238, 238);\n"
"border-bottom-right-radius: 50px;\n"
"border-top-right-radius: 50px;")
        self.label_3.setText("")
        self.label_3.setObjectName("label_3")
        self.comboBox_4 = QtWidgets.QComboBox(self.widget)
        self.comboBox_4.setGeometry(QtCore.QRect(1490, 370, 281, 41))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.comboBox_4.setFont(font)
        self.comboBox_4.setStyleSheet("background-color:rgba(0, 0, 0, 0);\n"
"border:none;\n"
"border-bottom:2px solid rgba(46, 82, 101, 200);\n"
"color:rgba(0, 0, 0, 240);\n"
"padding-bottom:7px;")
        self.comboBox_4.setObjectName("comboBox_4")
        self.comboBox_3 = QtWidgets.QComboBox(self.widget)
        self.comboBox_3.setGeometry(QtCore.QRect(1490, 220, 281, 41))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.comboBox_3.setFont(font)
        self.comboBox_3.setStyleSheet("background-color:rgba(0, 0, 0, 0);\n"
"border:none;\n"
"border-bottom:2px solid rgba(46, 82, 101, 200);\n"
"color:rgba(0, 0, 0, 240);\n"
"padding-bottom:7px;")
        self.comboBox_3.setObjectName("comboBox_3")
        self.comboBox_5 = QtWidgets.QComboBox(self.widget)
        self.comboBox_5.setGeometry(QtCore.QRect(1490, 300, 281, 41))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.comboBox_5.setFont(font)
        self.comboBox_5.setStyleSheet("background-color:rgba(0, 0, 0, 0);\n"
"border:none;\n"
"border-bottom:2px solid rgba(46, 82, 101, 200);\n"
"color:rgba(0, 0, 0, 240);\n"
"padding-bottom:7px;")
        self.comboBox_5.setObjectName("comboBox_5")
        self.comboBox_2 = QtWidgets.QComboBox(self.widget)
        self.comboBox_2.setGeometry(QtCore.QRect(1490, 150, 281, 41))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.comboBox_2.setFont(font)
        self.comboBox_2.setStyleSheet("background-color:rgba(0, 0, 0, 0);\n"
"border:none;\n"
"border-bottom:2px solid rgba(46, 82, 101, 200);\n"
"color:rgba(0, 0, 0, 240);\n"
"padding-bottom:7px;")
        self.comboBox_2.setObjectName("comboBox_2")
        self.lineEdit_5 = QtWidgets.QLineEdit(self.widget)
        self.lineEdit_5.setGeometry(QtCore.QRect(1490, 80, 281, 41))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.lineEdit_5.setFont(font)
        self.lineEdit_5.setStyleSheet("background-color:rgba(0, 0, 0, 0);\n"
"border:none;\n"
"border-bottom:2px solid rgba(46, 82, 101, 200);\n"
"color:rgba(0, 0, 0, 240);\n"
"padding-bottom:7px;")
        self.lineEdit_5.setObjectName("lineEdit_5")
        self.pushButton_14 = QtWidgets.QPushButton(self.widget)
        self.pushButton_14.setGeometry(QtCore.QRect(1750, 50, 101, 51))
        font = QtGui.QFont()
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_14.setFont(font)
        self.pushButton_14.setStyleSheet("QPushButton{\n"
"    background-color: qlineargradient(spread:pad, x1:0, y1:0.505682, x2:1, y2:0.477, stop:0 rgba(11, 131, 120, 219), stop:1 rgba(85, 98, 112, 226));\n"
"    color:rgba(255, 255, 255, 210);\n"
"    border-top-right-radius: 50px;\n"
"}\n"
"\n"
"QPushButton:hover{\n"
"    background-color: qlineargradient(spread:pad, x1:0, y1:0.505682, x2:1, y2:0.477, stop:0 rgba(150, 123, 111, 219), stop:1 rgba(85, 81, 84, 226));\n"
"}\n"
"\n"
"QPushButton:pressed{\n"
"    padding-left:5px;\n"
"    padding-top:5px;\n"
"    background-color:rgba(150, 123, 111, 255);\n"
"}\n"
"\n"
"\n"
"\n"
"border-left-radius: 25px;")
        icon5 = QtGui.QIcon()
        icon5.addPixmap(QtGui.QPixmap(":/newPrefix/Stok/search.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.pushButton_14.setIcon(icon5)
        self.pushButton_14.setIconSize(QtCore.QSize(27, 27))
        self.pushButton_14.setObjectName("pushButton_14")
        self.pushButton_14.clicked.connect(self.sorgudetay)
        self.retranslateUi(Form)
        QtCore.QMetaObject.connectSlotsByName(Form)


    def retranslateUi(self, Form):
        _translate = QtCore.QCoreApplication.translate
        Form.setWindowTitle(_translate("Form", "LİSTELE"))
        item = self.tableWidget.horizontalHeaderItem(0)
        item.setText(_translate("Form", "Kayıt Tarihi"))
        item = self.tableWidget.horizontalHeaderItem(1)
        item.setText(_translate("Form", "Stok Kodu"))
        item = self.tableWidget.horizontalHeaderItem(2)
        item.setText(_translate("Form", "Stok Adı"))
        item = self.tableWidget.horizontalHeaderItem(3)
        item.setText(_translate("Form", "Birim"))
        item = self.tableWidget.horizontalHeaderItem(4)
        item.setText(_translate("Form", "Adet"))
        item = self.tableWidget.horizontalHeaderItem(5)
        item.setText(_translate("Form", "Gurubu"))
        item = self.tableWidget.horizontalHeaderItem(6)
        item.setText(_translate("Form", "Ara Gurubu"))
        item = self.tableWidget.horizontalHeaderItem(7)
        item.setText(_translate("Form", "Marka"))
        item = self.tableWidget.horizontalHeaderItem(8)
        item.setText(_translate("Form", "Ürün Cinsi"))
        item = self.tableWidget.horizontalHeaderItem(9)
        item.setText(_translate("Form", "Model"))
        item = self.tableWidget.horizontalHeaderItem(10)
        islem.execute("Select * From a")
        self.a = islem.fetchall()
        for ia in self.a:
            self.comboBox_2.addItem(ia[0])
        islem.execute("Select * From b")
        self.b = islem.fetchall()
        for ib in self.b:
            self.comboBox_3.addItem(ib[0])
        islem.execute("Select * From c")
        self.c = islem.fetchall()
        for ic in self.c:
            self.comboBox_4.addItem(ic[0])
        islem.execute("Select * From d")
        self.d = islem.fetchall()
        for id in self.d:
            self.comboBox_5.addItem(id[0])
        sorgu = "select * from ana"
        for eb, r in enumerate(islem.execute(sorgu)):
            for d, i in enumerate(r):
                self.tableWidget.setItem(eb, d, QtWidgets.QTableWidgetItem(i))
        __sortingEnabled = self.tableWidget.isSortingEnabled()
        self.tableWidget.setSortingEnabled(False)
        self.tableWidget.setSortingEnabled(__sortingEnabled)
        self.pushButton_9.setText(_translate("Form", "SİL"))
        self.pushButton_10.setText(_translate("Form", "DÜZENLE"))
        self.pushButton_11.setText(_translate("Form", "KAPAT"))
        self.pushButton_12.setText(_translate("Form", "YAZDIR"))
        self.pushButton_13.setText(_translate("Form", "EXCEL"))
        self.lineEdit_5.setPlaceholderText(_translate("Form", "Tarih Giriniz."))
        self.pushButton_14.setText(_translate("Form", "ARA"))
    def satirsil(self):
        self.secili = self.tableWidget.currentRow()
        self.secili_id1 = self.tableWidget.item(self.secili, 1).text()
        self.tableWidget.removeRow(self.secili)
        islem.execute("DELETE FROM ana WHERE stokk='{}'".format(self.secili_id1))
        baglanti.commit()
    def excel_dok(self):
        self.dosya = Workbook()
        self.sayfa = self.dosya.active
        self.sayfa.title = "Ürün Listesi"
        self.sayfa["A1"].value = "Kayıt Tarihi"
        self.sayfa["B1"].value = "Stok Kodu"
        self.sayfa["C1"].value = "Stok Adı"
        self.sayfa["D1"].value = "Birim"
        self.sayfa["E1"].value = "Adet"
        self.sayfa["F1"].value = "Gurubu"
        self.sayfa["G1"].value = "Ara Gurubu"
        self.sayfa["H1"].value = "Marka"
        self.sayfa["I1"].value = "Model"
        self.sayfa["J1"].value = "Ürün Cinsi"
        self.sayfa["K1"].value = "Açıklama"
        self.dosya.save("./ürünlistesi.xlsx")
        self.dosya_ac = openpyxl.load_workbook("./ürünlistesi.xlsx")
        self.sayfa_ac = self.dosya_ac["Ürün Listesi"]
        self.sayfa_sayisi = self.sayfa_ac.max_row
        islem.execute("SELECT * FROM ana")
        self.islem = islem.fetchall()
        self.toplam = 2
        self.yatay = 1
        for i in self.islem:
            self.sayfa_ac.cell(self.toplam, self.yatay, value="{}".format(i[0]))
            self.yatay += 1
            self.sayfa_ac.cell(self.toplam, self.yatay, value="{}".format(i[1]))
            self.yatay += 1
            self.sayfa_ac.cell(self.toplam, self.yatay, value="{}".format(i[2]))
            self.yatay += 1
            self.sayfa_ac.cell(self.toplam, self.yatay, value="{}".format(i[3]))
            self.yatay += 1
            self.sayfa_ac.cell(self.toplam, self.yatay, value="{}".format(i[4]))
            self.yatay += 1
            self.sayfa_ac.cell(self.toplam, self.yatay, value="{}".format(i[5]))
            self.yatay += 1
            self.sayfa_ac.cell(self.toplam, self.yatay, value="{}".format(i[6]))
            self.yatay += 1
            self.sayfa_ac.cell(self.toplam, self.yatay, value="{}".format(i[7]))
            self.yatay += 1
            self.sayfa_ac.cell(self.toplam, self.yatay, value="{}".format(i[8]))
            self.yatay += 1
            self.sayfa_ac.cell(self.toplam, self.yatay, value="{}".format(i[9]))
            self.yatay += 1
            self.sayfa_ac.cell(self.toplam, self.yatay, value="{}".format(i[10]))
            self.toplam += 1
            self.yatay -= 10
        self.dosya_ac.save("./ürünlistesi.xlsx")

    def hucre(self):
        icerik = ui3.tableWidget.currentItem().text()
        secilisatir = ui3.tableWidget.currentRow()
        secilikolon = ui3.tableWidget.currentColumn()
        id = ui3.tableWidget.item(secilisatir, 1).text()
        if secilikolon == 0:
            islem.execute("UPDATE ana SET tarih='{}' Where stokk='{}'".format(icerik,id))
        elif secilikolon == 2:
            islem.execute("UPDATE ana SET stoka='{}' Where stokk='{}'".format(icerik,id))
        elif secilikolon == 3:
            islem.execute("UPDATE ana SET birim='{}' where stokk='{}'".format(icerik,id))
        elif secilikolon == 4:
            islem.execute("UPDATE ana SET adet='{}' where stokk='{}'".format(icerik,id))
        elif secilikolon == 5:
            islem.execute("UPDATE ana SET gurubu='{}' where stokk='{}'".format(icerik,id))
        elif secilikolon == 6:
            islem.execute("UPDATE ana SET arag='{}' where stokk='{}'".format(icerik,id))
        elif secilikolon == 7:
            islem.execute("UPDATE ana SET marka='{}' where stokk='{}'".format(icerik,id))
        elif secilikolon == 8:
            islem.execute("UPDATE ana SET cinsi='{}' where stokk='{}'".format(icerik,id))
        elif secilikolon == 9:
            islem.execute("UPDATE ana SET model='{}' where stokk='{}'".format(icerik,id))
        elif secilikolon == 10:
            islem.execute("UPDATE ana SET aciklama='{}' where stokk='{}'".format(icerik,id))
        baglanti.commit()
    def sorgudetay(self):
        self.tableWidget.clear()
        if self.comboBox_5.currentText() != "Marka Seçiniz" and self.comboBox_5.currentText() != "Ürün Seçiniz" and self.comboBox_3.currentText() != "Ara Gurubu Seçiniz" and self.comboBox_2.currentText() != "Gurup Seçiniz" and len(self.lineEdit_5.text()) != 0:
            self.tarih = self.lineEdit_5.text()
            self.gurup = self.comboBox_2.currentText()
            self.ara = self.comboBox_3.currentText()
            self.urun = self.comboBox_5.currentText()
            self.marka = self.comboBox_4.currentText()
            sorgu = "select * from ana where tarih='{}' and gurubu='{}' and arag='{}' and cinsi ='{}' and marka = '{}'".format(self.tarih,self.gurup,self.ara,self.urun,self.marka)
            for eb, r in enumerate(islem.execute(sorgu)):
                for d, i in enumerate(r):
                    self.tableWidget.setItem(eb, d, QtWidgets.QTableWidgetItem(i))
        elif self.comboBox_2.currentText() != "Gurup Seçiniz":
            self.gurup =self.comboBox_2.currentText()
            sorgu = "select * from ana where gurubu='{}'".format(self.gurup)
            for eb, r in enumerate(islem.execute(sorgu)):
                for d, i in enumerate(r):
                    self.tableWidget.setItem(eb, d, QtWidgets.QTableWidgetItem(i))

        elif self.comboBox_3.currentText() != "Ara Gurubunu Seçiniz":
            self.ara = self.comboBox_3.currentText()
            sorgu = "select * from ana where arag='{}'".format(self.ara)
            for eb, r in enumerate(islem.execute(sorgu)):
                for d, i in enumerate(r):
                    self.tableWidget.setItem(eb, d, QtWidgets.QTableWidgetItem(i))
        elif self.comboBox_5.currentText() != "Ürün Seçiniz":
            self.urun = self.comboBox_5.currentText()
            sorgu = "select * from ana where cinsi='{}'".format(self.urun)
            for eb, r in enumerate(islem.execute(sorgu)):
                for d, i in enumerate(r):
                    self.tableWidget.setItem(eb, d, QtWidgets.QTableWidgetItem(i))
        elif self.comboBox_4.currentText() != "Marka Seçiniz":
            self.marka = self.comboBox_4.currentText()
            sorgu = "select * from ana where marka='{}'".format(self.marka)
            for eb, r in enumerate(islem.execute(sorgu)):
                for d, i in enumerate(r):
                    self.tableWidget.setItem(eb, d, QtWidgets.QTableWidgetItem(i))
        elif len(self.lineEdit_5.text()) != 0:
            self.tarih = self.lineEdit_5.text()
            sorgu = "select * from ana where tarih='{}'".format(self.tarih)
            for eb, r in enumerate(islem.execute(sorgu)):
                for d, i in enumerate(r):
                    self.tableWidget.setItem(eb, d, QtWidgets.QTableWidgetItem(i))
    def print(self):
     os.startfile("ürünlistesi.xlsx","print")
import resim1_rc
if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    listeleme_pencere = QtWidgets.QMainWindow()
    ui3 = Ui_Form_listele()
    ui3.setupUi(listeleme_pencere)
    listeleme_pencere.show()
    ui3.tableWidget.cellChanged.connect(Ui_Form_listele.hucre)
    app.exec_()
